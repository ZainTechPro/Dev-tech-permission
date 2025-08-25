
import openpyxl
from openpyxl.styles import Protection

# Load the workbook and select the active sheet
workbook = openpyxl.load_workbook("Permissions_Matrix.xlsx")
sheet = workbook.active

# Define permission columns (assuming they start from column C, i.e., index 3)
permission_columns = ["Read Access", "Write Access", "Modify Access", "Delete Access", "Full Control"]

# Find the column indices for permission types
header = [cell.value for cell in sheet[1]]
permission_col_indices = [header.index(col) + 1 for col in permission_columns]

# Add data validation for checkboxes (TRUE/FALSE) and unlock cells
for col_idx in permission_col_indices:
    for row_idx in range(2, sheet.max_row + 1):  # Start from row 2 (after header)
        cell = sheet.cell(row=row_idx, column=col_idx)
        cell.data_type = openpyxl.cell.cell.TYPE_BOOL # Set data type to boolean
        cell.value = False # Default to False
        cell.protection = Protection(locked=False) # Unlock the cell

# Protect the sheet, allowing only unlocked cells to be edited
sheet.protection.sheet = True
sheet.protection.autoFilter = False
sheet.protection.sort = False
sheet.protection.pivotTable = False
sheet.protection.objects = False
sheet.protection.scenarios = False

# Save the modified workbook
workbook.save("Permissions_Matrix_with_Features.xlsx")

print("Permissions_Matrix_with_Features.xlsx created with checkboxes and protection.")


